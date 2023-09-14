import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension
import os

class ExcelTaskManager:
    def __init__(self, filename):
        self.filename = filename
        self.workbook = None
        self.sheet = None

    def open_excel_file(self):
        if os.path.isfile(self.filename):
            self.workbook = openpyxl.load_workbook(self.filename)
            self.sheet = self.workbook.active
        else: self.create_excel_file()

    def create_excel_file(self):
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = "Task Manager"
        headers = ["Task", "Description", "Date", "Time", "Urls"]
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            cell = self.sheet.cell(row=1, column=col_num, value=header)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            self.sheet.column_dimensions[col_letter] = ColumnDimension(self.sheet, min=15, max=30)

    def write_data(self, data):
        if not self.workbook:
            self.open_excel_file()
        
        next_row = self.sheet.max_row + 1
        self.sheet.append(data)

    def check_excel_file_existence(self):
        return os.path.isfile(self.filename)

    def remove_data(self, task_name):
        if not self.workbook:
            return
        
        rows_to_delete = []
        for row_idx, row in enumerate(self.sheet.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] == task_name:
                rows_to_delete.append(row_idx)
        for row_idx in reversed(rows_to_delete):
            self.sheet.delete_rows(row_idx)
        # self.sheet.reset_rows()
        
    def read_all_rows(self):
        if self.check_excel_file_existence():
            if not self.workbook:
                self.open_excel_file()
            rows = [list(row) for row in self.sheet.iter_rows(min_row=2, values_only=True)]
            return rows
        else: return []

    def save_excel_file(self):
        if self.workbook:
            self.workbook.save(self.filename)

if __name__ == "__main__":
    excel_filename = "./resources/Tasks.xlsx"
    excel_manager = ExcelTaskManager(excel_filename)
    # sample_data = ["Task 1", "Description 1", "2023-09-10", "09:00 AM", "https://example.com"]
    excel_manager.open_excel_file()
    excel_manager.remove_data("gym5")
    excel_manager.save_excel_file()
    # excel_manager.write_data(sample_data)
    # print(excel_manager.read_all_rows())
    # excel_manager.save_excel_file()

